import sqlite3
from openpyxl import Workbook
import os
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

CAMINHO_BANCO = os.path.join(os.path.dirname(__file__), 'usuarios.db')


def criar_tabela():
    conexao = sqlite3.connect(CAMINHO_BANCO)
    cursor = conexao.cursor()
    cursor.execute(''' 
        CREATE TABLE IF NOT EXISTS usuarios(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            telefone TEXT NOT NULL,
            cpf TEXT NOT NULL UNIQUE,
            data_nascimento TEXT NOT NULL,
            status INTEGER DEFAULT 1
        )
    ''')
    conexao.commit()
    conexao.close()


def inserir_usuario(nome, email, telefone, cpf, data_nascimento):
    try:
        conexao = sqlite3.connect(CAMINHO_BANCO)
        cursor = conexao.cursor()
        cursor.execute(''' 
            INSERT INTO usuarios (nome, email, telefone, cpf, data_nascimento, status)
            VALUES (?, ?, ?, ?, ?, 1)
        ''', (nome, email, telefone, cpf, data_nascimento))
        conexao.commit()
        return 'Usuário cadastrado com sucesso!'

    except sqlite3.IntegrityError as e:
        if 'email' in str(e):
            return 'Erro: Email já cadastrado!'
        elif 'cpf' in str(e):
            return 'Erro: CPF já cadastrado!'
        else:
            return 'Dados duplicados ou inválidos!'

    finally:
        conexao.close()


def listar_usuarios():
    conexao = sqlite3.connect(CAMINHO_BANCO)
    cursor = conexao.cursor()
    cursor.execute('SELECT * FROM usuarios')
    usuarios = cursor.fetchall()
    conexao.close()
    return usuarios


def editar_usuario():
    conexao = sqlite3.connect(CAMINHO_BANCO)
    cursor = conexao.cursor()
    cursor.execute('SELECT * FROM usuarios')
    usuarios = cursor.fetchall()

    if not usuarios:
        print('Nenhum usuário encontrado!')
        conexao.close()
        return

    print('\nUsuários cadastrados:')
    for u in usuarios:
        status = 'Ativo' if u[6] == 1 else 'Inativo'
        print(f'ID: {u[0]} | Nome: {u[1]} | Email: {u[2]} | Telefone: {u[3]} | CPF: {u[4]} | Data Nasc.: {u[5]} | Status: {status}')

    try:
        id_usuario = int(input('\nDigite o ID do usuário que deseja editar: '))
    except ValueError:
        print('ID inválido!')
        conexao.close()
        return

    cursor.execute('SELECT * FROM usuarios WHERE id = ?', (id_usuario,))
    usuario = cursor.fetchone()

    if not usuario:
        print('Usuário não encontrado!')
        conexao.close()
        return

    print('\nDeixe o campo em branco para manter o valor atual:')
    novo_nome = input(f'Nome ({usuario[1]}): ') or usuario[1]
    novo_email = input(f'Email ({usuario[2]}): ') or usuario[2]
    novo_telefone = input(f'Telefone ({usuario[3]}): ') or usuario[3]
    novo_cpf = input(f'CPF ({usuario[4]}): ') or usuario[4]
    nova_data = input(f'Data de Nascimento ({usuario[5]}): ') or usuario[5]

    while True:
        entrada_status = input(f'Status [1=Ativo / 0=Inativo] ({usuario[6]}): ')
        if entrada_status == '':
            novo_status = usuario[6]
            break
        elif entrada_status in ['0', '1']:
            novo_status = int(entrada_status)
            break
        else:
            print('Valor inválido! Digite apenas 1 para ativo ou 0 para inativo.')


    cursor.execute('SELECT id FROM usuarios WHERE email = ? AND id != ?', (novo_email, id_usuario))
    if cursor.fetchone():
        print('Erro: O Email já está em uso! Digite outro email.')
        conexao.close()
        return
    
    cursor.execute('SELECT id FROM usuarios WHERE cpf = ? AND id != ?', (novo_cpf, id_usuario))
    if cursor.fetchone():
        print('Erro: O CPF já está em uso! Digite outro CPF.')
        conexao.close()
        return
    
    while True:
        print('\nConfirme as alterações:')
        print(f'1. Nome: {novo_nome}')
        print(f'2. Email: {novo_email}')
        print(f'3. Telefone: {novo_telefone}')
        print(f'4. CPF: {novo_cpf}')
        print(f'5. Data de Nascimento: {nova_data}')
        print(f'6. Status: {"Ativo" if novo_status == 1 else "Inativo"}')

        confirm = input('Digite enter para salvar todas as alterações ou digite o número da opção que deseja alterar novamente: ').strip().lower()
        if confirm == '':
            break
        elif confirm == '1':
            novo_nome = input(f'Novo nome ({novo_nome}): ') or novo_nome
        elif confirm == '2':
            novo_email = input(f'Novo email ({novo_email}): ') or novo_email
        elif confirm == '3':
            novo_telefone = input(f'Novo telefone ({novo_telefone}): ') or novo_telefone
        elif confirm == '4':
            novo_cpf = input(f'Novo CPF ({novo_cpf}): ') or novo_cpf
        elif confirm == '5':
            nova_data = input(f'Nova data de Nascimento ({nova_data}): ') or nova_data
        elif confirm == '6':
            while True:
                entrada_status = input(f'Status [1=Ativo / 0=Inativo] ({novo_status}): ')
                if entrada_status == '':
                    novo_status = usuario[6]
                    break
                elif entrada_status in ['0', '1']:
                    novo_status = int(entrada_status)
                    break
                else:
                    print('Valor inválido! Digite apenas 1 para ativo ou 0 para inativo.')
        
    

    cursor.execute('''
        UPDATE usuarios
        SET nome = ?, email = ?, telefone = ?, cpf = ?, data_nascimento = ?, status = ?
        WHERE id = ?
    ''', (novo_nome, novo_email, novo_telefone, novo_cpf, nova_data, int(novo_status), id_usuario))
    conexao.commit()
    conexao.close()
    print('\nUsuário atualizado com sucesso!')


def remover_usuario():
    conexao = sqlite3.connect(CAMINHO_BANCO)
    cursor = conexao.cursor()
    cursor.execute('SELECT * FROM usuarios')

    usuarios = cursor.fetchall()

    if not usuarios:
        print('Usuário não encontrado')
        conexao.close()
        return
    
    print('\nUsuários cadastrados:')
    for u in usuarios:
        status = 'Ativo' if u[6] == 1 else 'Inativo'
        print(f'ID: {u[0]} | Nome: {u[1]} | Email: {u[2]} | Telefone: {u[3]} | CPF: {u[4]} | Data Nasc.: {u[5]} | Status: {status}')

        try:
            id_usuario = int(input('\nDigite o ID do usuário que deseja tornar inativo: '))
        except ValueError:
            print('ID inválido! Insira um ID válido.')
            conexao.close()
            return
    
    cursor.execute('SELECT * FROM usuarios WHERE id = ?', (id_usuario,))
    usuario = cursor.fetchone()

    if not usuario:
        print('Usuário não encontrado!')
        conexao.close()
        return
    
    if usuario[6] == 0:
        print('Esse usuário já se encontra inativo ')
        conexao.close()
        return
    
    confirmar = input(f'Tem certeza que deseja tornar o usuario "{usuario[1]}" inativo? (s/n)').strip().lower()
    if confirmar != 's':
        print('Operação cancelada')
        conexao.close()
        return
    
    cursor.execute('UPDATE usuarios SET status = 0 WHERE id = ?', (id_usuario,))
    conexao.commit()
    conexao.close()
    
    print(f'Usuário "{usuario[1]}" desativado com sucesso!')

def reativar_usuario():
    conexao = sqlite3.connect(CAMINHO_BANCO)
    cursor = conexao.cursor()
    cursor.execute('SELECT * FROM usuarios')
    usuarios = cursor.fetchall()

    if not usuarios:
        print('Usuário não encontrado')
        conexao.close()
        return
    
    print('\nUsuários cadastrados:')
    for u in usuarios:
        status = 'Inativo' if u[6] == 0 else 'Ativo'
        print(f'ID: {u[0]} | Nome: {u[1]} | Email: {u[2]} | Telefone: {u[3]} | CPF: {u[4]} | Data Nasc.: {u[5]} | Status: {status}')

    try:
        id_usuario = int(input('\nDigite o ID do usuário que deseja tornar ativo: '))
    except ValueError:
        print('ID inválido! Insira um ID válido.')
        conexao.close()
        return
    
    cursor.execute('SELECT * FROM usuarios WHERE id = ?', (id_usuario,))
    usuario = cursor.fetchone()

    if not usuario:
        print('Usuário não encontrado!')
        conexao.close()
        return
    
    if usuario[6] == 1:
        print('Esse usuário já se encontra ativo ')
        conexao.close()
        return
    
    confirmar = input(f'Tem certeza que deseja tornar o usuario "{usuario[1]}" ativo? (s/n)').strip().lower()
    if confirmar != 's':
        print('Operação cancelada')
        conexao.close()
        return
    
    cursor.execute('UPDATE usuarios SET status = 1 WHERE id = ?', (id_usuario,))
    conexao.commit()
    conexao.close()
    
    print(f'Usuário "{usuario[1]}" ativado com sucesso!')



def exportar_excel():
    print("\n1. Exportar apenas usuários ativos")
    print("2. Exportar todos os usuários")

    escolha = input('Digite sua escolha: ')
    
    
    conexao = sqlite3.connect(CAMINHO_BANCO)
    cursor = conexao.cursor()

    if escolha == '1':
        cursor.execute('SELECT * FROM usuarios WHERE status = 1')
        nome_arquivo = 'clientes_ativos.xlsx'
        titulo = 'Clientes ativos'
    
    elif escolha == '2':
        cursor.execute('SELECT * FROM usuarios')
        nome_arquivo = 'Clientes_todos.xlsx'
        titulo = 'Lista de clientes'
    
    else: 
        print('Opção inválida! Tente novamente. ')
        conexao.close()
        return
    
    usuarios = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = titulo

    cabecalhos = ['ID', 'Nome', 'Email', 'Telefone', 'CPF', 'Data de Nascimento', 'Status']
    sheet.append(cabecalhos)

    for u in usuarios:
        status = 'Ativo' if u[6] == 1 else 'Inativo'
        linha = [u[0], u[1], u[2], u[3], u[4], u[5], status]
        sheet.append(linha)

    
    # --- ESTILOS ---
    header_font = Font(bold=True, color="FFFFFF")  # branco
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # azul suave
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Aplica estilo ao cabeçalho
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center

    # Ajusta largura automática das colunas
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width

    caminho_excel = os.path.join(os.path.dirname(__file__), nome_arquivo)
    workbook.save(caminho_excel)
    conexao.close()
    print(f'Tabela exportada com sucesso! Arquivo salvo em: "{caminho_excel}"')

